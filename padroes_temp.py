import openpyxl
import os

planTemp = None
sheetPlanTemp = None

def temp():
    file_path = "pasta\\temp.xlsx"

    if os.path.exists(file_path):
        planTemp = openpyxl.load_workbook(file_path)
        sheetPlanTemp = planTemp.active
        print(f"Arquivo temp carregado com sucesso!")
    else:
        planTemp = openpyxl.Workbook()
        sheetPlanTemp = planTemp.active

        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        planTemp.save(file_path)
        print(f"Arquivo temp criado com sucesso!")

    return planTemp, sheetPlanTemp

planTemp, sheetPlanTemp = temp()

def adicionar_info(informacoes):

    linhas = informacoes.split('\n')

    planTemp = openpyxl.load_workbook("pasta\\temp.xlsx")
    sheetPlanTemp = planTemp.active


    proxima_linha = sheetPlanTemp.max_row + 1


    for linha in linhas:

        dados = linha.split("\t")  
        for j, valor in enumerate(dados, start=1):
            sheetPlanTemp.cell(row=proxima_linha, column=j).value = valor
        proxima_linha += 1


    # Salvar as alterações na planilha
    planTemp.save('pasta\\temp.xlsx')
